"""This is a python program to scrap the NPD online excel sheet
and automatically generate delay reports """
# Written by E.Amir Anwar
import os
import platform
import sys
from time import strftime

import colorama
from numpy import busday_count
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from prettytable import PrettyTable
from termcolor import cprint, colored

colorama.init()
# initial configuration for the windows console to prevent recieving character encoding errors
my_path = os.path.dirname(os.path.realpath(sys.argv[0]))
system32_path = os.path.join(os.environ['SystemRoot'],
                             'SysNative' if platform.architecture()[0] == '32bit' else 'System32')
os.chdir(system32_path)
os.system('chcp 65001')
os.system('set PYTHONIOENCODING=utf-8')
os.chdir(my_path)

# global variables:
book = None
sheet_names = None
# fields to be used for varios types of reports
regular_fields = [r'Code', r'Description', r'Set type', r'SITUATION', r'ORDER NO.', r'EXCEPTED IN WAREHOUSE',
                  r'RECVD AT WAREHOUSE']
payment_fields = ['Code', 'Description', 'Set type', 'Supplier', 'SITUATION', 'ORDER NO.', 'Order Date']
trial_delays = ['Code', 'Description', 'SITUATION', 'PR Date', 'Order Date', 'Tech.Approval for Manufac.',
                'EXCEPTED IN WAREHOUSE']
full_delays = ['Code', 'Description', 'SITUATION', 'PR Date', 'Order Date', 'Tech.Approval for Manufac.',
               'EXCEPTED IN WAREHOUSE']
pending_trials = ['Code', 'Description', 'SITUATION', 'ORDER NO.', 'RECVD AT WAREHOUSE']
all_reports = [regular_fields, payment_fields, trial_delays, full_delays, pending_trials]
month_dict = {'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04', 'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
              'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'}


def month_name_to_number(date):
    month = date.split('-')[1]
    try:
        int(month)

    except:
        month_number = month_dict[month.lower()]
        date = date.replace(month, month_number)
    return date


def dif_two_dates(date1, date2, max_diff):
    try:
        formatted_d1 = (str(date1)).split(' ')[0]
        formatted_d2 = (str(date2)).split(' ')[0]
        if formatted_d2 == 'None':
            formatted_d2 = (lambda x: strftime("%Y-%m-%d"))('None')
        formatted_d1 = month_name_to_number(formatted_d1)
        formatted_d2 = month_name_to_number(formatted_d2)
        difference = busday_count(formatted_d1, formatted_d2, weekmask="Sun Mon Tue Wed Thu")
        if difference > max_diff:
            return True
        else:
            return False
    except:
        pass


def data_to_table(li, field_number):
    table = PrettyTable(hrules=1, vrules=1, border=1, field_names=li[0], header=True, encoding='UTF-8')
    for i in range(1, len(li)):
        if field_number == '1':
            table.add_row(li[i])
        elif field_number == '2':
            if li[i][10] == 'PAYMENT':
                table.add_row(li[i])
        elif field_number == '3':
            isTrial = li[i][6] == 'NPD TRIAL'
            if isTrial:
                isNotRecieved = (li[i][10] != 'RECVD AT WH') and (li[i][10] != 'CANCELLED') and (
                    li[i][10] != 'FINISHED') and (li[i][10] != 'SALE HOLD')
                is_technical_delayed = dif_two_dates(li[i][14], li[i][18], 15)
                is_delivery_delayed = dif_two_dates(li[i][18], li[i][20], 28)
                is_total_delayed = dif_two_dates(li[i][14], li[i][20], 43)
                if (isTrial and isNotRecieved) and (is_technical_delayed or is_delivery_delayed or is_total_delayed):
                    table.add_row(li[i])
        elif field_number == '4':
            isFull = li[i][6] == 'NPD FULL' or 'Reorder'
            if isFull:
                isNotRecieved = (li[i][10] != 'RECVD AT WH') and (li[i][10] != 'CANCELLED') and (
                    li[i][10] != 'FINISHED')
                is_technical_delayed = dif_two_dates(li[i][14], li[i][18], 15)
                is_delivery_delayed = dif_two_dates(li[i][18], li[i][20], 28)
                is_total_delayed = dif_two_dates(li[i][14], li[i][20], 43)
                if (isFull and isNotRecieved) and (is_technical_delayed or is_delivery_delayed or is_total_delayed):
                    table.add_row(li[i])
        elif field_number == '5':
            isTrial = li[i][6] == 'NPD TRIAL'
            if isTrial:
                isRecieved = (li[i][10] == 'RECVD AT WH')
                isNotTrialed = li[i][22] == 'None' or '' or ' ' or None
                if isTrial and isRecieved and isNotTrialed:
                    table.add_row(li[i])

    return table


def get_fields(report_number):
    if report_number == '1':
        return regular_fields
    elif report_number == '2':
        return payment_fields
    elif report_number == '3':
        return trial_delays
    elif report_number == '4':
        return full_delays
    elif report_number == '5':
        return pending_trials
    elif report_number == '6':
        return all_reports
    else:
        new_input = input('Unrecognized choice, please enter a number from 1 to 6\n')
        get_fields(new_input)


def sheet_to_list(work_sheet):
    # getting a generator
    generator = work_sheet.iter_rows()
    # getting the list of tuples from the generator
    li = []
    for tu in generator:  # type: tuple
        li.append(list(tu))  # adding the tuple to l after converting it to a list 
    for l in li:  # replacing every cell object with its value
        for c in l:  # type: Cell
            cell_index = l.index(c)
            val = c.value
            l[cell_index] = val
            del c  # removing the cell from memory
    return li


def get_worksheet_obj(sheet_name):
    for name in sheet_names:
        if name == sheet_name:
            return book.get_sheet_by_name(name)
    print('Sheet name was not found !')


def print_sheet_names():
    sheets_string = ''
    for i in sheet_names:
        sheets_string += ", " + i
    sheets_string = sheets_string[2:]  # removing the extra , and space in the beginning
    cprint(sheets_string, 'white', 'on_green')


def main():
    # constant strings:
    report_types = ' 1- All Data, 2- Payment Issues, 3- Trial set delays, 4- Full set delays, ' \
                   '5- Trials received but not done, ' \
                   '6- Generate all reports without viewing(recommended)'
    report_types_list = report_types.split(',')
    warning_message = r"Warning: some reports might not display correctly because of encoding issues, " \
                      r"So it is recommended that you generate all reports to disc directly (option no. 6)"
    print('Choose one of the following sheets:')
    print_sheet_names()
    user_sheet = input('Please enter the sheet name you wish to view\n')
    sheet = get_worksheet_obj(user_sheet)
    print('Please wait...')
    l = sheet_to_list(sheet)
    del sheet
    cprint(warning_message, 'white', 'on_red')
    report_number = input(
        "Which type of report you wish to generate:\n" + report_types + "\n Please choose a number.. ")
    fields = get_fields(report_number)
    if report_number != '6':
        report_name = report_types_list[int(report_number) - 1][3:]
        generate_report(fields, l, report_name, report_number, user_sheet)
    elif report_number == '6':
        make_all_reports(fields, l, report_types_list, user_sheet)


def make_all_reports(fields, l, report_types_list, user_sheet):
    count = 0
    for i in fields:
        t = data_to_table(l, str(count + 1))
        table_string = u'{x}'.format(x=t.get_string(fields=i))
        file_name = (user_sheet + '{x}'.format(x=report_types_list[int(count)][3:]) + '.txt')  # type : str
        with open(file_name, 'w', errors='ignore') as f:
            f.write(table_string)
        count += 1


def generate_report(fields, li, report_name, report_number, user_sheet):
    t = data_to_table(li, report_number)
    del li  # saving memory by removing the big list l
    print(colored('Table Title: ' + user_sheet, 'white', 'on_blue'))
    table_string = t.get_string(fields=fields)
    print(table_string, flush=True)
    save_choice = input('Do you want to save this table? (y/n)')
    if save_choice == 'y':
        file_name = (user_sheet + report_name + '.txt')  # type : str
        with open(file_name, 'w', errors='ignore') as f:
            f.write(table_string)


def intro_dialog():
    cprint('Welcome to the NPD auto report generation tool', 'white', 'on_blue')
    cprint('This program was written by E. Amir Anwar on 02-04-2017', 'white', 'on_blue')
    input('Please make sure that the NPD.xlsm file is in the same folder with this program, then press enter..')
    print('Please wait..')
    path = os.getcwd() + r'\NPD.xlsm'
    global book
    global sheet_names
    book = load_workbook(path)
    sheet_names = book.sheetnames
    print('Successfully opened the workbook')


# Program Starting
intro_dialog()
while True:
    try:
        main()

    except:
        cprint('An Error Happened!!', 'white', 'on_red')
        main()
